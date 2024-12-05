# -*- coding: utf-8 -*-
"""
Created on Mon Apr 17 12:43:11 2023

@author: jsavage
"""

import sys
import pandas as pd
import sympy as sym
import numpy as np
import math as math
import scipy
import statsmodels.api as sm
import matplotlib.pyplot as plt
import xlsxwriter
import statsmodels.formula.api as smf
from statsmodels.iolib.summary2 import summary_col
from openpyxl import Workbook
import openpyxl
from openpyxl.chart import BarChart, Reference, Series

x_import = r"F:\DerekS\My Projects\Python\Meeder1\Residual Xs.xlsx"
y_import = r"F:\DerekS\My Projects\Python\Meeder1\10 Year Price Return Ys.xlsx"

#For Winsorization
upper_limit = 1
lower_limit = 0

deciles = 5

lag = 2

#To specify the time frame of forward returns you want to analyze. 'd' is daily, 'w' is weekly, and 'm' is monthly. Letters are lowercase.
forward_returns = 'w'

#To smooth out the x factor
x_roll = 0

print(x_roll)
print(forward_returns)


############################# Dont Change Anything Under Here Unless You Know Python ############################################
workbook = Workbook()
wb = openpyxl.load_workbook(r'F:\DerekS\My Projects\Python\Meeder1\Quantile Output.xlsx')
sheet_names = wb.sheetnames

for name in sheet_names:
    sheet = wb[name]
    wb.remove(sheet)
new_sheet = wb.create_sheet('Sheet1')
wb.active = new_sheet

wb.save(r'F:\DerekS\My Projects\Python\Meeder1\Quantile Output.xlsx')


try:    
    df_x = pd.read_excel(x_import)
    df_x['Date'] = pd.to_datetime(df_x['Date'])
    df_x.set_index('Date', inplace = True)
    df_x = df_x.sort_index()
    df_x = df_x.iloc[:,:].shift(lag)
    
    df_y = pd.read_excel(y_import)
    df_y['Date'] = pd.to_datetime(df_y['Date'])
    df_y.set_index('Date', inplace = True)
    df_y = df_y.sort_index()
except ValueError: 
    df_x = pd.read_csv(x_import, sep = '\t')
    df_x['Date'] = pd.to_datetime(df_x['Date'])
    df_x.set_index('Date', inplace = True)
    df_x = df_x.sort_index()
    df_x = df_x.iloc[:,:].shift(lag)
    
    df_y = pd.read_csv(y_import, sep = '\t')
    df_y['Date'] = pd.to_datetime(df_y['Date'])
    df_y.set_index('Date', inplace = True)
    df_y = df_y.sort_index()
else:
    df_x = pd.read_excel(x_import)
    df_x['Date'] = pd.to_datetime(df_x['Date'])
    df_x.set_index('Date', inplace = True)
    df_x = df_x.sort_index()
    df_x = df_x.iloc[:,:].shift(lag)
    
    df_y = pd.read_excel(y_import)
    df_y['Date'] = pd.to_datetime(df_y['Date'])
    df_y.set_index('Date', inplace = True) 
    df_y = df_y.sort_index()    
        
df_x_raw = df_x    

    
df_xy = pd.merge(df_y, df_x_raw, on='Date')
df_xy = df_xy.sort_index()

df_decile = pd.DataFrame()
summary_df = pd.DataFrame()

x_factors = df_xy.iloc[:,1:]
y = df_xy.iloc[:,0]
y_factor = y.name
first_date = df_xy.index.min()
first_date = first_date.strftime('%Y-%m-%d')
last_date = df_xy.index.max()
last_date = last_date.strftime('%Y-%m-%d')

reg_dict = {}
rsquared_dict = {}
plots = []
results_df = pd.DataFrame()
decile_dict = {}
pvalue_dict = {}
df_fake = pd.DataFrame()

for factor in df_xy:
    try:
        df_reg = df_xy[[y_factor, factor]]
        df_fake[factor + 'Decile'] = pd.qcut(df_reg[factor], deciles, labels=False)
    except ValueError:
        continue
    else:
            if factor == y_factor:
                continue
            df_reg = pd.DataFrame()
            df_reg = df_xy[[y_factor, factor]]
            df_reg = df_reg.dropna()
            if forward_returns == 'd':
                pass
            elif forward_returns == 'w':
                df_reg[factor] = df_reg[factor].rolling(window=x_roll).mean()
                df_reg[y_factor] = df_y.rolling(window=5).sum().shift(-5)
                df_reg = df_reg.resample('W').last()
            elif forward_returns == 'm':
                df_reg[factor] = df_reg[factor].rolling(window=x_roll).mean()
                df_reg[y_factor] = df_y.rolling(window=21).sum().shift(-21)
                df_reg = df_reg.resample('M').last()
            else:
                print('Please provide a valid forward_return time frame')
                sys.exit()
            reg_upperlimit = df_reg[factor].quantile(upper_limit)
            reg_lowerlimit = df_reg[factor].quantile(lower_limit)
            df_reg['winsorized ' + factor] = df_reg[factor].clip(reg_lowerlimit, reg_upperlimit)
            df_reg = df_reg.dropna()
            y = df_reg.iloc[:,0]
            x = df_reg[factor]
            x = sm.add_constant(x)
            reg = sm.OLS(y,x).fit()
            reg_dict[factor] = reg
            rsquared_dict[factor] = reg.rsquared
            pvalue_dict[factor] = reg.pvalues[1]
            df_decile = pd.DataFrame()
            df_decile[factor + 'Decile'] = pd.qcut(df_reg[factor], deciles, labels=False)
            df_decile['Returns'] = y
            avg_returns = df_decile.groupby(factor + 'Decile')["Returns"].mean()
            decile_dict[factor] = avg_returns
            n = df_decile.shape[0] / deciles
            tstat_calc = (df_decile.groupby(factor + 'Decile')["Returns"].mean() / df_decile.groupby(factor + 'Decile')["Returns"].std()) * math.sqrt(n)

        
         # write the deciles to a new sheet in the Excel file
            results_dict = {'R-squared': reg.rsquared, 'p-value': reg.pvalues[1]}
            factor_sheet_name = factor[:29].replace('/','')
            new_sheet_name = wb.create_sheet(factor_sheet_name)
            new_sheet_name.cell(row=1, column=1, value = factor)
            new_sheet_name.cell(row=3, column=1, value = 'Deciles')
            for d in range(1,1+deciles):
                new_sheet_name.cell(row=d+3, column=1, value=d)
            for index, value in avg_returns.items():
                index = int(index)
                new_sheet_name.cell(row=index+4, column=2, value=round(value,4))
            i = 0
            for tstat in tstat_calc:
                new_sheet_name.cell(row=i+4, column=3, value=round(tstat,4))
                i = i+1
            sheet = wb[factor_sheet_name]
            data = Reference(sheet, min_col=2, min_row=4, max_col=2, max_row= 3 + deciles)
            chart = BarChart()
            chart.add_data(data)
            chart.title = factor + " Decile Analysis"
            chart.x_axis.title = 'Deciles'
            chart.y_axis.title = 'Returns'
            sheet.add_chart(chart, 'E7')
            new_sheet_name.cell(row=1, column=4, value = 'R-squared')
            new_sheet_name.cell(row=2, column=4,value = round(results_dict['R-squared'],4))
            new_sheet_name.cell(row=1, column=5,value = 'p-value')
            new_sheet_name.cell(row=2, column=5,value = round(results_dict['p-value'],4))

x_factors_differenced = pd.DataFrame()
positive_check = (x_factors.dropna() >= 0).all()
for col in x_factors:
    if positive_check[col]:
        x_factors_differenced[col] = df_x[col].pct_change()
    if not positive_check[col]:
        x_factors_differenced[col] = df_x[col].diff()
        

        
word = 'Differenced '
x_factors_differenced.columns = [word + col for col in x_factors_differenced.columns]
y = df_xy.iloc[:,0]

df_xy_diff = pd.merge(y, x_factors_differenced, on='Date')
y = df_xy_diff.iloc[:,0]
#x_factors_differenced = df_xy_diff.iloc[:,1:]
df_decile_diff = pd.DataFrame()

for factor in df_xy_diff:
    try:
        df_fake[factor + 'Decile'] = pd.qcut(df_xy_diff[factor], deciles, labels=False)
    except ValueError:
        continue
    else:
       if factor == y_factor:
            continue
       df_reg = df_xy_diff[[y_factor, factor]]
       df_reg = df_reg.dropna()
       if forward_returns == 'd':
           pass
       elif forward_returns == 'w':
           df_reg = df_reg.rolling(window=x_roll).mean()
           df_reg[y_factor] = df_y.rolling(window=5).sum().shift(-5)
           df_reg = df_reg.resample('W').last()
       elif forward_returns == 'm':
           df_reg = df_reg.rolling(window=x_roll).mean()
           df_reg[y_factor] = df_y.rolling(window=21).sum().shift(-21)
           df_reg = df_reg.resample('M').last()
       reg_upperlimit = df_reg[factor].quantile(upper_limit)
       reg_lowerlimit = df_reg[factor].quantile(lower_limit)
       df_reg['winsorized ' + factor] = df_reg[factor].clip(reg_lowerlimit, reg_upperlimit)
       y = df_reg.iloc[:,0]
       y = y.dropna()
       x = df_reg[factor]
       x = x.dropna()
       x = sm.add_constant(x)
       align = pd.merge(x,y, on='Date')
       x = align.iloc[:,:2]
       y = align.iloc[:,2]
       reg = sm.OLS(y,x).fit()
       reg_dict[factor] = reg
       rsquared_dict[factor] = reg.rsquared
       pvalue_dict[factor] = reg.pvalues[1]
       df_decile_diff = pd.DataFrame()
       df_decile_diff[factor + 'Decile'] = pd.qcut(df_reg[factor], deciles, labels=False)
       df_decile_diff['Returns'] = y
       avg_returns_differenced = df_decile_diff.groupby(factor + 'Decile')["Returns"].mean()
       decile_dict[factor] = avg_returns_differenced
       n = df_decile_diff.shape[0] / deciles
       tstat_diff = (df_decile_diff.groupby(factor + 'Decile')["Returns"].mean() / df_decile_diff.groupby(factor + 'Decile')["Returns"].std()) * math.sqrt(n)
       
       results_dict = {'R-squared': reg.rsquared, 'p-value': reg.pvalues[1]}
       factor_sheet_name = factor[:29].replace('/','')
       new_sheet_name = wb.create_sheet(factor_sheet_name)
       new_sheet_name.cell(row=1, column=1, value = factor)
       new_sheet_name.cell(row=3, column=1, value = 'Deciles')
       for d in range(1,1+deciles):
           new_sheet_name.cell(row=d+3, column=1, value=d)
       for index, value in avg_returns_differenced.items():
           index = int(index)
           new_sheet_name.cell(row=index+4, column=2, value=round(value,4))
       i = 0
       for tstat in tstat_diff:
           new_sheet_name.cell(row=i+4, column=3, value=round(tstat,4))
           i = i+1
       sheet = wb[factor_sheet_name]
       data = Reference(sheet, min_col=2, min_row=4, max_col=2, max_row= 3 + deciles)
       chart = BarChart()
       chart.add_data(data)
       chart.title = factor + " Decile Analysis"
       chart.x_axis.title = 'Deciles'
       chart.y_axis.title = 'Returns'
       sheet.add_chart(chart, 'E7')
       new_sheet_name.cell(row=1, column=4, value = 'R-squared')
       new_sheet_name.cell(row=2, column=4,value = round(results_dict['R-squared'],4))
       new_sheet_name.cell(row=1, column=5,value = 'p-value')
       new_sheet_name.cell(row=2, column=5,value = round(results_dict['p-value'],4))

sheet = wb["Sheet1"]
sheet.cell(row = 1, column=5, value = 'R-Squared and Decile Summary for Factors against ' + y_factor)
date_range = str(first_date) + ' to ' + str(last_date)
sheet.cell(row = 2, column = 5, value = date_range)
counter = 1
sheet.cell(row=1,column=1,value='Factors')
sheet.cell(row=1,column=2,value='R-squared')
sheet.cell(row=1,column=3,value='P-Values')
for key, value in rsquared_dict.items():
     counter = counter+1
     sheet.cell(row=counter+1, column=1, value=key)
counter=1
for key, value in rsquared_dict.items():
    counter = counter+1
    sheet.cell(row=counter+1, column=2, value=round(value,4))
counter=1
for key, value in pvalue_dict.items():
    counter = counter+1
    sheet.cell(row=counter+1, column=3, value=round(value,2))   
counter=1
for d in range(1,deciles+1):
    sheet.cell(row=4+d, column=5, value = d)
for key, value in decile_dict.items():
    counter = counter + 1
    sheet.cell(row=4, column=4+counter, value = key)
counter = 1
for key, value in decile_dict.items():
    dict_counter = 0 
    counter = counter + 1
    for c in range(1,deciles+1):
        sheet.cell(row=4+c, column=4+counter, value = value[dict_counter])
        dict_counter = dict_counter + 1
        if dict_counter == deciles:
            break   
        
df_correlation = pd.merge(x_factors, x_factors_differenced, on='Date')
correlation_matrix = df_correlation.corr()

wb.save(r'F:\DerekS\My Projects\Python\Meeder1\Quantile Output.xlsx')
wb.close()

# writer = pd.ExcelWriter("F:\DerekS\My Projects\Python\Meeder1\Correlation Matrix.xlsx" ,engine= 'xlsxwriter')
# workbook = writer.book

# worksheet = workbook.add_worksheet("Matrix")
# writer.sheets["Matrix"] = worksheet
# correlation_matrix.to_excel(writer,sheet_name="Matrix",startrow=0,startcol=0)
# writer.save()